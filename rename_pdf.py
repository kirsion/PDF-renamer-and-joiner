import fitz
import os
from PyPDF2 import PdfFileMerger, PdfFileReader
from os import DirEntry, curdir, getcwd, chdir, rename
from glob import glob as glob
import time
import openpyxl
start_time = time.time()



directory = 'PDF_FILES'
curr_dir = getcwd()

chdir(directory)

pdf_list = glob('*.pdf')


# reads excel
path = r"C:\Users\kirsi\Documents\python\pdf\CBT AccountsTEST.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['CBT Acct#']

exc_acc_num = []
exc_prop_name = []

for i in range(1, 450, 1):
    try:
        acc_num_list = exc_acc_num.append(sheet.cell(row=i+1, column=5).value)
        prop_name_list = exc_prop_name.append(sheet.cell(row=i+1, column=4).value)
    except:
        pass

exc_acc_num[:] = (elem[:10] for elem in exc_acc_num if elem is not None)

for pdf in pdf_list:
    #print(pdf)
    with fitz.open(pdf) as pdf_obj:
        text = pdf_obj[0].get_text()
        #print(text)
    accname = text.split("\n")[5].strip()
    trunc_accname = accname[8:]
    datename = text.split("\n")[3].strip()
    trunc_datename = datename[16:]
    year = trunc_datename[-2:]
    month = trunc_datename[:3]

    trunc_accname_list = list()
    trunc_accname_list.append(trunc_accname)

    if any(x in trunc_accname_list for x in exc_acc_num):
        indexlist = []
        i = 0
        while (i < len(exc_acc_num)):
            if (trunc_accname_list.count(exc_acc_num[i]) > 0):
                indexlist.append(i)
            i += 1
        for i in range(0, len(indexlist)):
            indexlist_int = int(indexlist[i])
        pdf_prop_name = exc_prop_name[indexlist_int]

    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(1)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(2)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(3)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(4)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(5)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(6)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(7)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(8)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(9)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(10)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError

pdfs = {}
for f in glob('*.pdf'):
    name = f[:f.find('(')] + f[f.find(')') + 1:]
    pdfs.setdefault(name, []).append(f)

for outfile, group in pdfs.items():
    merge = PdfFileMerger()
    for x in group:
      try:
        merge.append(x)
      except KeyError:
        print('Unable to merge:', x)
    merge.write(outfile)
    merge.close()

pdf_list = glob('*.pdf')
for pdf in pdf_list:
    if '(' in pdf:
        os.remove(pdf)
    else:
        pass


print("--- %s seconds ---" % (time.time() - start_time))













