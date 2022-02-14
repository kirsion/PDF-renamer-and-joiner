#pymupdf
import fitz
import os
from PyPDF2 import PdfFileMerger
from os import getcwd, chdir, rename
from glob import glob as glob
import time
import openpyxl

#gets the runtime
start_time = time.time()

#sets the directory to where the pdf files to be renamed and merged will be
directory = 'PDF_FILES'
curr_dir = getcwd()
chdir(directory)

#creates a list of pdf files in the 'PDF_FILES' folder
pdf_list = glob('*.pdf')

#finds the path of the excel file for renaming the pdf files
path = r"C:\Users\vpham\PycharmProjects\pythonProject\CBT AccountsTEST.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['CBT Acct#']


exc_acc_num = []
exc_prop_name = []
#creates lists for the needed information in the excel file, in the column for each row
for i in range(1, 500, 1):
    try:
        acc_num_list = exc_acc_num.append(sheet.cell(row=i+1, column=5).value)
        prop_name_list = exc_prop_name.append(sheet.cell(row=i+1, column=4).value)
    except:
        pass
#truncates the excel account # list elements to only the numbers
exc_acc_num[:] = (elem[:10] for elem in exc_acc_num if elem is not None)

for pdf in pdf_list:
    with fitz.open(pdf) as pdf_obj:
        #extracts text from opened pdf file
        text = pdf_obj[0].get_text()
    #find the text where the account and date name is and splits it there
    accname = text.split("\n")[5].strip()
    trunc_accname = accname[8:]
    datename = text.split("\n")[3].strip()
    #Then truncates the string into the correct length
    trunc_datename = datename[16:]
    year = trunc_datename[-2:]
    month = trunc_datename[:3]
    #create list of the truncated account names
    trunc_accname_list = list()
    trunc_accname_list.append(trunc_accname)
    #finds if the name in pdf list is in the excel account number list
    if any(x in trunc_accname_list for x in exc_acc_num):
        indexlist = []
        i = 0
        while (i < len(exc_acc_num)):
            if (trunc_accname_list.count(exc_acc_num[i]) > 0):
                indexlist.append(i)
            i += 1
        for i in range(0, len(indexlist)):
            indexlist_int = int(indexlist[i])
            #links the index of the list in the trunc account name to the property name in the excel
        pdf_prop_name = exc_prop_name[indexlist_int]

    try: #renames the pdf file with the correct name.
         #But there are multiple files with same name so the duplicates are made with (#)
         #I hard coded the different iretations because I didn't know how to loop it. It works perfectly
         #if you run it once, but if you run it twice, the (#)'s get mixed around a bit, but if you run it
         #a second time, the ordering is reordered numerically.
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(1)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(2)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(3)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(4)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(5)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(6)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(7)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(8)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(9)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(10)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(11)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
    try:
        while not FileExistsError:
            rename(pdf, pdf_prop_name + '-Bkstmnt-' + month + year + '.pdf')
        else:
            rename(pdf, pdf_prop_name + '(12)' + '-Bkstmnt-' + month + year + '.pdf')
    except:
        FileExistsError
#creating dictionary to group similarly named flies with only (#) difference in filename


pdfs = {}
for f in glob('*.pdf'):
    #creates the final filename
    name = f[:f.find('(')] + f[f.find(')') + 1:]
    #creating the dictionary with the final filename as the key or index and the list of duplicate file names (1),(2),...
    pdfs.setdefault(name, []).append(f)
#merges each files in the group
for outfile, group in pdfs.items():
    merge = PdfFileMerger()
    for x in group:
      try:
        merge.append(x)
      except KeyError:
        print('Unable to merge:', x)
    merge.write(outfile)
    merge.close()
#deletes the duplicate pdfs
pdf_list = glob('*.pdf')
for pdf in pdf_list:
    if '(' in pdf:
        os.remove(pdf)
    else:
        pass

print("--- %s seconds ---" % (time.time() - start_time))













